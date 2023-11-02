import openpyxl
import os

def main():
    # Check if the expense report file exists
    if os.path.exists("informe_gastos.xlsx"):
        # If it exists, load the workbook and select the "Gastos" worksheet
        wb = openpyxl.load_workbook("informe_gastos.xlsx")
        ws = wb["Gastos"]
        next_row = ws.max_row + 1
    else:
        # If it doesn't exist, create a new workbook and select the "Gastos" worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gastos"
        next_row = 1
        # Write the headers to the worksheet
        ws.cell(row=next_row, column=1, value="Fecha")
        ws.cell(row=next_row, column=2, value="Descripción")
        ws.cell(row=next_row, column=3, value="Monto")
        next_row += 1

    # Prompt the user to enter details of each expense
    expenses = []
    while True:
        date = input("Ingrese la fecha del gasto (en formato dd/mm/aaaa): ")
        description = input("Ingrese una descripción del gasto: ")
        amount = input("Ingrese el monto del gasto: ")
        try:
            amount = float(amount)
        except ValueError:
            print("El monto debe ser un número.")
            continue
        expenses.append((date, description, amount))
        if input("¿Desea ingresar otro gasto? (s/n): ").lower() != "s":
            break

    # Write the expenses to the worksheet
    for row, expense in enumerate(expenses, start=next_row):
        ws.cell(row=row, column=1, value=expense[0])
        ws.cell(row=row, column=2, value=expense[1])
        ws.cell(row=row, column=3, value=expense[2])

    # Calculate the total expenses and find the most expensive and cheapest expenses
    total = sum(expense[2] for expense in expenses)
    most_expensive = max(expenses, key=lambda x: x[2])
    cheapest = min(expenses, key=lambda x: x[2])

    # Print the expense summary
    print(f"Total de gastos: {total}")
    print(f"Gasto más caro: {most_expensive[0]} - {most_expensive[1]} (${most_expensive[2]})")
    print(f"Gasto más barato: {cheapest[0]} - {cheapest[1]} (${cheapest[2]})")

    # Save the workbook and print a message indicating completion
    wb.save("informe_gastos.xlsx")
    print("El informe de gastos ha sido generado y guardado en el archivo informe_gastos.xlsx.")

if __name__ == "__main__":
    main()